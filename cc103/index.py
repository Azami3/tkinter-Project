from tkinter import *
import customtkinter
import tkinter as tk
from tkinter import ttk 
from tkinter.ttk import Progressbar
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = tk.Tk()
root.geometry("900x600")
root.configure(bg="pink")
root.title("MAINPAGE")
root.resizable(False,False)



class HPD:
    def __init__(self, root, num_pages, page_contents):
        self.root = root
        self.num_pages = num_pages
        self.page_contents = page_contents
        

        pagination_frame = tk.Frame(root, bg="#CD1076")
        pagination_frame.place(x=380, y=150)  # Adjust the position as needed

    def update_active_dot(self, active_page):
        for index, page_content in enumerate(self.page_contents):
            if index == active_page:
                page_content.place(x=305, y=48)
            else:
                page_content.pack_forget()
    


def login():
    
    root1 = Toplevel()
    root1.geometry("400x500")
    root1.configure(bg="pink")
    root1.resizable(False,False)
    root1.title("LOGIN/SIGNUP")
    
    def loginwindow():
        global center_frame,ter_frame,signupnoption,userLog,passLog
        ter_frame = customtkinter.CTkFrame(master=root1,
                                            height=500,
                                            width=400,
                                            bg_color="black",
                                            fg_color="pink",
                                            corner_radius=10,)
        ter_frame.place(x=1,y=1)

        logo2 = ImageTk.PhotoImage(Image.open("logo.png").resize((100, 100), Image.LANCZOS))
        logo_l = customtkinter.CTkLabel(master=ter_frame, image=logo2,text="")
        logo_l.place(x=300, y=1)
        center_frame = customtkinter.CTkFrame(master=ter_frame,
                                            height=410,
                                            width=410,
                                            bg_color="black",
                                            fg_color="black",
                                            
                                            )
        center_frame.place(x=1,y=100)
        w = customtkinter.CTkLabel(master=center_frame,text="Welcome to Game Garage",
                font=("Consolas",20,"bold"),text_color ="white")
        w.place(x=50,y=50)
        x = customtkinter.CTkLabel(master=center_frame,text="Login here",
                font=("Consolas",20,"normal"),text_color ="white")
        x.place(x=50,y=80)
        user = customtkinter.CTkLabel(master=center_frame,text="Username:",
                font=("Consolas",12,"normal"),text_color ="white")
        user.place(x=50,y=125)
        userLog=customtkinter.CTkEntry(master=center_frame, width=240,corner_radius=0,
                                        font=('Calibri Gothic',13), 
                                    placeholder_text=' Enter your username',
                                    border_color="#CD1076",border_width=0,
                                    placeholder_text_color="#ffffff",
                                    fg_color="grey",
                                    text_color="white",
                                    )
        userLog.place(x=50,y=150)
        Pass = customtkinter.CTkLabel(master=center_frame,text="Password:",
                font=("Consolas",12,"normal"),text_color ="white")
        Pass.place(x=50,y=185)
        passLog=customtkinter.CTkEntry(master=center_frame, width=240,corner_radius=0,
                                        font=('Calibri',13), 
                                    placeholder_text=' Enter your Password',
                                    border_color="#a9b388",border_width=0,
                                    placeholder_text_color="#ffffff",
                                    fg_color="grey",show="•",
                                    text_color="#ffffff",)
        passLog.place(x=50,y=210)
        loginBtn = customtkinter.CTkButton(master=center_frame,font=('Calibri',20,"bold"),
                                        width=120, 
                                        text="Login",
                                        border_color="#e09132",
                                        border_width=0,
                                        hover_color="#C4661F",
                                        fg_color="#e09132",
                                        text_color="white",
                                        command=lambda:enter_login()
                                        )
        loginBtn.place(x=50, y=300)
        def showpass():
            if showpass_var.get():
                passLog.configure(show="")
            else:
                passLog.configure(show="•")
        showpass_var = BooleanVar()
        showpass_var.set(False)
        show_pass = customtkinter.CTkCheckBox(master=center_frame,text="Show password",
                                            font=('consolas',12,),text_color="#ffffff",
                                            checkbox_height=12,checkbox_width=12,checkmark_color="#000000",
                                            fg_color="#FFFFFF",hover_color="#FFFFFF",
                                            border_width=1,border_color="#FFFFFF",variable=showpass_var,
                                            command=showpass
                                            )
        show_pass.place(x=50,y=240)
        def point1():
            global pointer1
            pointer1= customtkinter.CTkLabel(master=ter_frame,
                                            font=('Calibri',20,"bold"),
                                            height=1,width=4,
                                            text="|",
                                            text_color="#e09132")
            pointer1.place(x=40,y=50)
        def point2():
            global pointer2
            pointer2 = customtkinter.CTkLabel(master=ter_frame,
                                            font=('Calibri',20,"bold"),
                                            height=1,width=4,
                                            text="|",
                                            text_color="#e09132")
            pointer2.place(x=140,y=50)
        point1()
        
        def on_enter_login(event):
            loginoption.configure(font=("Calibri",20,"bold"),
                                cursor="hand2",)
            loginoption['foreground'] = '#1c1c1c'
        def on_leave_login(event):
            loginoption.configure(font=("Calibri",20,"normal"),
                                state=DISABLED)
        def on_click_login(event):
            loginoption.configure(font=("Calibri",20,"bold"))
            signupnoption.configure(font=("Calibri",20,"normal"))
            
            loginwindow()
        loginoption = customtkinter.CTkLabel(master=ter_frame,text="Login",
                            font=("Calibri",20,"bold"),state=NORMAL
                            )
        loginoption.place(x=50,y=50)
        loginoption.bind("<Enter>", on_enter_login)
        loginoption.bind("<Leave>", on_leave_login)
        loginoption.bind("<Button-1>", on_click_login)


        def on_enter_Signup(event):
            signupnoption.configure(font=("Calibri",20,"bold"),cursor="hand2")
        def on_leave_Signup(event):
            signupnoption.configure(font=("Calibri",20,"normal"))
        def on_click_Signup(event):
            point2()
            signupnoption.configure(font=("Calibri",20,"bold"),cursor="hand2")
            loginoption.configure(font=("Calibri",20,"normal"))
            pointer1.destroy()
            
            sign_up()


    

        signupnoption = customtkinter.CTkLabel(master=ter_frame,text="Sign up",
                            font=("Calibri",20,"normal"),
                            )
        signupnoption.place(x=150,y=50)
        signupnoption.bind("<Enter>", on_enter_Signup)
        signupnoption.bind("<Leave>", on_leave_Signup)
        signupnoption.bind("<Button-1>", on_click_Signup)
    loginwindow()
    root1.mainloop()


def fromSignintologin(event):
    signup.destroy()
    login()
def sign_up():
    center_frame.destroy()
    global signup,for_username,for_Contact,for_regpass
    signup = customtkinter.CTkFrame(master=ter_frame,
                                          height=410,
                                          width=410,
                                          #"#1c1c1c",
                                          )
    signup.place(x=1,y=100)

    w = customtkinter.CTkLabel(master=signup,text="Register here",
              font=("Consolas",25,"bold"))
    w.place(x=50,y=50)
    x = customtkinter.CTkLabel(master=signup,text="fill the Requirements",
              font=("Consolas",20,"normal"))
    x.place(x=50,y=80)
    user = customtkinter.CTkLabel(master=signup,text="Username:",
              font=("Consolas",12,"normal"))
    user.place(x=50,y=115)
    for_username=customtkinter.CTkEntry(master=signup, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                  placeholder_text=' Enter your name',
                                  border_color="#a9b388",border_width=0,
                                  placeholder_text_color="#ffffff",
                                  fg_color="grey",
                                  text_color="#ffffff",)
    for_username.place(x=50,y=140)
    Pass = customtkinter.CTkLabel(master=signup,text="Password:",
              font=("Consolas",12,"normal"))
    Pass.place(x=50,y=175)
    for_regpass=customtkinter.CTkEntry(master=signup, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                  placeholder_text=' Enter your desired Password',
                                  border_color="#a9b388",border_width=0,
                                  placeholder_text_color="#ffffff",
                                  fg_color="grey",
                                  text_color="#ffffff",)
    for_regpass.place(x=50,y=200)
    
    contact = customtkinter.CTkLabel(master=signup,text="Email Address:",
              font=("Consolas",12,"normal"))
    contact.place(x=50,y=235)
    for_Contact=customtkinter.CTkEntry(master=signup, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                  placeholder_text=' Enter your  Email Address.',
                                  border_color="#a9b388",border_width=0,
                                  placeholder_text_color="#ffffff",
                                  fg_color="grey",
                                  text_color="#ffffff",)
    for_Contact.place(x=50,y=260)
    regBtn = customtkinter.CTkButton(master=signup,font=('Calibri',20,"bold"),
                                       width=120, 
                                       text="Register",
                                       border_color="#e09132",
                                       border_width=0,
                                       hover_color="#C4661F",
                                       fg_color="#e09132",
                                       text_color="white",
                                       command=lambda:registration_Checker()
                                       )
    regBtn.place(x=50, y=330)
def registration_Checker():
    Excel = Workbook()
    Excel = load_workbook('login_signin.xlsx')
    Excel_Acc = Excel.active
    reg_username = for_username.get()
    same_user = False
    Reg_pass = for_username.get()
    confirm_pass = for_Contact.get()

    if for_username.get() == "" or for_regpass.get() == "" or for_Contact.get() == "" :
        messagebox.showerror("Entries", f'All entries must be filled')
    else:
        print("Entries passed")
        for each_cells in range(2, Excel_Acc.max_row + 1):
            if reg_username == Excel_Acc['A' + str(each_cells)].value:
                same_user = True
                break
        if same_user:
            messagebox.showerror("USERNAME", f'Username already taken.')
        else:
            add_row = str(Excel_Acc.max_row + 1)
            Excel_Acc['A' + str(add_row)] = for_username.get()        
            Excel_Acc['B' + str(add_row)] = for_regpass.get()
            Excel_Acc['C' + str(add_row)] = for_Contact.get()

            messagebox.showinfo("Regitered", f'Congrats you are registered')
            Excel.save('login_signin.xlsx')  
            Excel.close()
def user_passDel():
    userLog.delete(0,END)
    passLog.delete(0,END)
def enter_login():
    global root1
    Excel = Workbook()
    Excel = load_workbook('login_signin.xlsx')
    Excel_Acc = Excel.active
    user_entry = userLog.get()
    pass_entry = passLog.get()
    no_data = False
    if user_entry == "" and pass_entry == "":
        messagebox.showerror("DATA", f'Username or Password Fields are empty')
        no_data = True
    elif pass_entry == "":
        messagebox.showerror("DATA", f'Password Fields are empty')
        no_data = True
    elif user_entry == "" :
        messagebox.showerror("DATA", f'Username Fields are empty')
        no_data = True
    
    if user_entry == "" and pass_entry == "":
        messagebox.showerror("DATA", f'Username or Password Fields are empty')
        no_data = True
    elif pass_entry == "":
        messagebox.showerror("DATA", f'Password Fields are empty')
        no_data = True
    elif user_entry == "":
        messagebox.showerror("DATA", f'Username Fields are empty')
        no_data = True
    
    if not no_data:
        for each_cell in range(2, (Excel_Acc.max_row) + 1):
            if (user_entry == Excel_Acc['A' + str(each_cell)].value) and (pass_entry == Excel_Acc['B' + str(each_cell)].value):
                messagebox.showinfo('LOGIN', f'Login Successful')
                Excel.close()


                Mainpage()
                print("Login Successful")
                
                break
            elif (user_entry == Excel_Acc['A' + str(each_cell)].value):
                messagebox.showwarning('LOGIN', f'Password is incorrect')
                passLog.delete(0, END)
                break
        else:
            user_passDel()
            messagebox.showerror("LOGIN", f'Login Failed, Invalid Username or Password')
            print("Invalid User or Password")
        
        Excel.close()
    tk.mainloop()

def Mainpage():

    mainFrame = Frame(root,height=600,width=900, bg= "pink"
                      )
    mainFrame.place(x=0,y=0)
    loginBtn = customtkinter.CTkButton(master=mainFrame,font=('Calibri',20,"bold"),
                                       width=120, 
                                       text="sign up",
                                       border_color="#e09132",
                                       border_width=0,
                                       hover_color="#C4661F",
                                       fg_color="#e09132",
                                       text_color="white",
                                       command=lambda:signinnn()
                                       )
    loginBtn.place(x=700, y=10)

    def signinnn():
        login()
    logo = ImageTk.PhotoImage(Image.open("logo.png").resize((300, 100), Image.LANCZOS))
    logo_l = customtkinter.CTkLabel(master=mainFrame, image=logo,text="")
    logo_l.place(x=1, y=48)
    
    bg1 = ImageTk.PhotoImage(Image.open("bg1.png").resize((900, 450), Image.LANCZOS))
    bg_l = customtkinter.CTkLabel(root, image=bg1,text="")
    bg_l.place(x=1, y=150)

    greet = customtkinter.CTkLabel(master=mainFrame, text= "Welcome to Game Garage", 
                                   width=20, 
                                   font=("Cooper Black",25),
                                   bg_color="pink", 
                                   fg_color="pink")
    greet.place(x=240,y=10)

    num_pages = 100
    dot_images = [
        ImageTk.PhotoImage(Image.open("dot1.png").resize((590, 100), Image.LANCZOS)),
        ImageTk.PhotoImage(Image.open("dot2.png").resize((590, 100), Image.LANCZOS)),
        ImageTk.PhotoImage(Image.open("dot3.png").resize((590, 100), Image.LANCZOS))
     ]

    page_contents = [
        customtkinter.CTkLabel(root, image=image,text="") for image in dot_images * num_pages
    ]

    pagination = HPD(root, num_pages, page_contents)

    active_page = 0
    pagination.update_active_dot(active_page)

    def next_page():
        nonlocal active_page
        active_page = (active_page + 1) % (num_pages * len(dot_images))
        pagination.update_active_dot(active_page)
        root.after(2000, next_page)

    root.after(2000, next_page)

    mlbb1 = ImageTk.PhotoImage(Image.open("mlbb.png").resize((150, 150), Image.LANCZOS))
    mlbb_l = customtkinter.CTkButton(root, image=mlbb1,text="",
                                     border_color="black",
                                     bg_color = "black",
                                     fg_color ="black",
                                     command=lambda:ml_interface() 
                                     

                                     )
    mlbb_l.place(x=50, y=230)


    def ml_interface():
        root.withdraw()
        root2 = Toplevel()
        root2.geometry("900x600")
        root2.configure(bg="pink")
        root2.resizable(False,False)
        root2.title("MOBILE LEGENDS")

        def ml_to_main():
            root2.destroy()
            root.deiconify()

        back_button = customtkinter.CTkButton(master=root2, font=('Calibri', 20, "bold"),
                                                width=220, text="Back to Main Interface",
                                                border_color="#e09132",
                                                border_width=0,
                                                hover_color="#C4661F",
                                                fg_color="#e09132",
                                                text_color="white",
                                                command=lambda:ml_to_main()
                                                )
        back_button.place(x=650,y = 10)

        excel_con=Workbook()
        excel_con=load_workbook('ml_cart_items.xlsx')
        excel_activate=excel_con.active
        

        m_frame = customtkinter.CTkFrame(master=root2,
                                            height=750,
                                            width=390,
                                            bg_color="pink",
                                            fg_color="#CD1076",
                                            corner_radius=10,)
        m_frame.place(x=1,y=1)


        chartml = ImageTk.PhotoImage(Image.open("mlbbchart.jpg").resize((500, 200), Image.LANCZOS))
        chartml_l = customtkinter.CTkLabel(master=root2, image=chartml,text="")
        chartml_l.place(x=400, y=50)

        us_id = customtkinter.CTkLabel(master=m_frame, text="User ID:",
                                      font=("Consolas BOLD",20,"normal")
                                      )
        us_id.place(x=10,y=10)

        us_id_e=customtkinter.CTkEntry(master=m_frame, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                     placeholder_text=' Enter your  Ingame ID.',
                                     border_color="#a9b388",border_width=0,
                                     placeholder_text_color="#ffffff",
                                     fg_color="grey",
                                     text_color="#ffffff")
        us_id_e.place(x=120,y=10)

        server_id = customtkinter.CTkLabel(master=m_frame, text="Server ID:",
                                      font=("Consolas BOLD",20,"normal")
                                      )
        server_id.place(x=10,y=50)

        server_id_e=customtkinter.CTkEntry(master=m_frame, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                     placeholder_text=' Enter your  Server ID.',
                                     border_color="#a9b388",border_width=0,
                                     placeholder_text_color="#ffffff",
                                     fg_color="grey",
                                     text_color="#ffffff")
        server_id_e.place(x=120,y=50)

        recharge_id = customtkinter.CTkLabel(master=m_frame, text="Select Recharge:",
                                      font=("Consolas BOLD",20,"normal")
                                      )
        recharge_id.place(x=10,y=90)

        price_select =StringVar()

        o11 = Radiobutton(master=m_frame, text=("11 Diamonds"),bg="#CD1076",value= "11 Diamonds = 11 php", variable=price_select)
        o11.place(x=10,y=120)
        
        o22 = Radiobutton(master=m_frame, text=("22 Diamonds"),bg="#CD1076", value= "22 Diamonds = 20 php", variable=price_select)
        o22.place(x=10,y=150)
        
        o56 = Radiobutton(master=m_frame, text=("56 Diamonds"),bg="#CD1076", value= "56 Diamonds = 48 php", variable=price_select)
        o56.place(x=10,y=180)
        
        o112 = Radiobutton(master=m_frame, text=("112 Diamonds"),bg="#CD1076", value= "112 Diamonds = 93 php", variable=price_select)
        o112.place(x=10,y=210)
        
        o168 = Radiobutton(master=m_frame, text=("168 Diamonds"),bg="#CD1076", value= "168 Diamonds = 140 php", variable=price_select)
        o168.place(x=10,y=240)
        
        o223 = Radiobutton(master=m_frame, text=("223 Diamonds"),bg="#CD1076",value= "223 Diamonds = 186 php", variable=price_select)
        o223.place(x=10,y=270)
        
        o279 = Radiobutton(master=m_frame, text=("279 Diamonds"),bg="#CD1076", value= "279 Diamonds = 233 php", variable=price_select)
        o279.place(x=10,y=300)
        
        o336 = Radiobutton(master=m_frame, text=("336 Diamonds"),bg="#CD1076", value= "336 Diamonds = 279 php", variable=price_select)
        o336.place(x=150,y=120)

        o402 = Radiobutton(master=m_frame, text=("402 Diamonds"),bg="#CD1076", value= "402 Diamonds = 337 php", variable=price_select)
        o402.place(x=150,y=150)

        o570 = Radiobutton(master=m_frame, text=("570 Diamonds"),bg="#CD1076", value= "570 Diamonds = 465 php", variable=price_select)
        o570.place(x=150,y=180)

        o804 = Radiobutton(master=m_frame, text=("804 Diamonds"),bg="#CD1076", value= "804 Diamonds = 662 php", variable=price_select)
        o804.place(x=150,y=210)

        o1163 = Radiobutton(master=m_frame, text=("1163 Diamonds"),bg="#CD1076", value= "1163 Diamonds = 920 php", variable=price_select)
        o1163.place(x=150,y=240)

        o2398 = Radiobutton(master=m_frame, text=("2398 Diamonds"),bg="#CD1076", value= "2398 Diamonds = 1840 php", variable=price_select)
        o2398.place(x=150,y=270)

        o6042 = Radiobutton(master=m_frame, text=("6042 Diamonds"),bg="#CD1076", value= "6042 Diamonds = 4550 php", variable=price_select)
        o6042.place(x=150,y=300)

        payment_id = customtkinter.CTkLabel(master=m_frame, text="Select Payment:",font=("Consolas BOLD",20,"normal"))
        payment_id.place(x=10, y= 330)

        payment_option = StringVar()

        load = Radiobutton(master=m_frame, text=("Load "),bg="#CD1076", value= "Load", variable=payment_option)
        load.place(x=10,y=360)

        gcash = Radiobutton(master=m_frame, text=("Gcash "),bg="#CD1076", value= "Gcash", variable=payment_option)
        gcash.place(x=10,y=390)

        paymaya = Radiobutton(master=m_frame, text=("Pay Maya "),bg="#CD1076", value= "Pay Maya", variable=payment_option)
        paymaya.place(x=130,y=360)

        Paypal = Radiobutton(master=m_frame, text=("Paypal "),bg="#CD1076", value= "Paypal", variable=payment_option)
        Paypal.place(x=130,y=390)

        bank_account = Radiobutton(master=m_frame, text=("Bank Account "),bg="#CD1076", value= "Bank Account", variable=payment_option)
        bank_account.place(x=240,y=360)

        payment_id = customtkinter.CTkLabel(master=m_frame, text="Enter Your Payment Account Here:",font=("Consolas BOLD",15,"normal"))
        payment_id.place(x=10, y= 420)

        input_payment=customtkinter.CTkEntry(master=m_frame, width=240,corner_radius=0,
                                     font=('Century Gothic',13), 
                                     placeholder_text=' Enter your Payment Acc...',
                                     border_color="#a9b388",border_width=0,
                                     placeholder_text_color="#ffffff",
                                     fg_color="grey",
                                     text_color="#ffffff")
        input_payment.place(x=10,y=450)

        excel_con=Workbook()
        excel_con=load_workbook('ml_cart_items.xlsx')
        excel_activate=excel_con.active
        
        save_button = customtkinter.CTkButton(master=m_frame,font=('Calibri',20,"bold"),
                                        width=120, 
                                        text="add to Cart",
                                        border_color="#e09132",
                                        border_width=0,
                                        hover_color="#C4661F",
                                        fg_color="#e09132",
                                        text_color="white",
                                        command=lambda:save_function()
                                        )
        save_button.place(x=10,y=500)

        delete_button = customtkinter.CTkButton(master=m_frame,font=('Calibri',20,"bold"),
                                        width=120, 
                                        text="Cancel Order",
                                        border_color="#e09132",
                                        border_width=0,
                                        hover_color="#C4661F",
                                        fg_color="#e09132",
                                        text_color="white",
                                        command=lambda:delete_function()
                                        )
        delete_button.place(x=150,y=500)

        complete_button = customtkinter.CTkButton(master=root2, font=('Calibri', 20, "bold"),
                                              width=220, text="Complete Purchase",
                                              border_color="#e09132",
                                              border_width=0,
                                              hover_color="#C4661F",
                                              fg_color="#e09132",
                                              text_color="white",
                                              command=lambda:complete_purchase())
        complete_button.place(x=650,y=550)
        



        def save_function():
            user = us_id_e.get()
            server = server_id_e.get()
            price  = price_select.get()
            payment = payment_option.get()
            in_payment = input_payment.get()

            for e_row in range(2, excel_activate.max_row + 1):
                    if user == excel_activate['A' + str(e_row)].value:
                        break
            addrow = str(excel_activate.max_row +1)
            excel_activate['A' + addrow] = user
            excel_activate['B' + addrow] = server
            excel_activate['C' + addrow] = price
            excel_activate['D' + addrow] = payment
            excel_activate['E' + addrow] = in_payment
            messagebox.showinfo("NOTIFICATION", "YOUR ORDER HAS BEEN ADDED TO CART")
                    
            refresh_data(tv1)
            excel_con.save("ml_cart_items.xlsx")

        def delete_function():
            excel_activate.delete_rows(2, excel_activate.max_row)
            excel_con.save("ml_cart_items.xlsx")
            refresh_data(tv1)
            messagebox.showinfo("NOTIFICATION", "All Order has been deleted.")

            
        def t_view():
            global tv1 

            someframe = Frame(master=root2,width=500, height=200, bg="gray")
            someframe.place(x= 400, y= 280)

            style = ttk.Style()
            style.theme_create("colored", parent="alt", settings={
                "Treeview": {
                    "configure": {"background": "pink", "fieldbackground": "#CD1076"}
                }
            })
            style.theme_use("colored")

            tv1 = ttk.Treeview(someframe, height=10)
            yscroll = Scrollbar(someframe, orient='vertical', command=tv1.yview)
            xscroll = Scrollbar(someframe, orient='horizontal', command=tv1.xview)
            tv1.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
            xscroll.pack(side ="bottom",fill ="x")
            yscroll.pack(side ="right",fill ="y")
            tv1['columns'] = ('User ID', 'Server ID', 'Order','Payment Method','Payment acc')
            tv1.column('#0', width=0)
            tv1.column('User ID', width=95)
            tv1.column('Server ID', width=95)
            tv1.column('Order', width=95)
            tv1.column('Payment Method', width=95)
            tv1.column('Payment acc', width=95)

            tv1.heading('#0', text='', anchor=W)
            tv1.heading('User ID', text='User ID', anchor=W)
            tv1.heading('Server ID', text='Server ID', anchor=CENTER)
            tv1.heading('Order', text='Order', anchor=W)
            tv1.heading('Payment Method', text='Payment Method', anchor=W)
            tv1.heading('Payment acc', text='Payment acc', anchor=W)

            
            for e_cell in range(2,(excel_activate.max_row)+1):
                tv1.insert(parent ='',index="end",text=str(e_cell),values=(excel_activate['A'+ str(e_cell)].value,
                                                                        excel_activate['B'+ str(e_cell)].value,
                                                                        ))
    
            ur_cart = customtkinter.CTkLabel(master=root2, text="YOUR CART:",
                                      font=("Consolas BOLD",20,"normal"),fg_color ="pink"
                                      )
            ur_cart.place(x=500,y=250)


            

            tv1.pack()
        t_view()


        def refresh_data(tree): 
            tree.delete(*tree.get_children())

            data = update()
            for item in data:
                tree.insert('', 'end', values=item)
        
        def update():
            update_value = list()
            for each_cell in range(2,(excel_activate.max_row)+1):
                update_value.append([excel_activate['A'+str(each_cell)].value,
                                    excel_activate['B'+str(each_cell)].value, 
                                    excel_activate['C'+str(each_cell)].value, 
                                    excel_activate['D'+str(each_cell)].value, 
                                    excel_activate['E'+str(each_cell)].value, ])
            return update_value

        def complete_purchase():
            selected_user = us_id_e.get()

            if not selected_user:
                messagebox.showerror("Error", "Please enter User ID.")
                return
            confirmation = messagebox.askyesno("Confirmation", "Are you sure you want to complete this purchase?")

            if confirmation:
                for e_row in range(2, excel_activate.max_row + 1):
                    if selected_user == excel_activate['A' + str(e_row)].value:
                        excel_activate.delete_rows(e_row, e_row)
                        break

                excel_con.save("ml_cart_items.xlsx")
                messagebox
                refresh_data(tv1)
            
            success_window = Toplevel()
            success_window.title("Success")
            success_window.geometry("400x400")
            
            greet_image = Image.open("greet.jpg").resize((400, 400), Image.LANCZOS)
            greet_image = ImageTk.PhotoImage(greet_image)
            greet_label = tk.Label(success_window, image=greet_image)
            greet_label.image = greet_image  # Store reference to prevent garbage collection
            greet_label.place(x=1,y=1)
        
            ok_button = tk.Button(success_window, text="OK",font=("arial",15),bg="#e09123",fg="white", command=success_window.destroy)
            ok_button.place(x=150,y=320)
            
            us_id_e.delete(0, END) 
            server_id_e.delete(0, END) 
            price_select.set("") 
            payment_option.set("")  
            input_payment.delete(0, END)  

            
            refresh_data(tv1)
        root2.mainloop()

    codm1 = ImageTk.PhotoImage(Image.open("codm.png").resize((150, 150), Image.LANCZOS))
    codm_l = customtkinter.CTkButton(root, image=codm1,text="",
                                     border_color="black",
                                     bg_color = "black",
                                     fg_color ="black",
                                     command=lambda:cod_interface() 
                                     
                                     )
    codm_l.place(x=350, y=230)


    def cod_interface():
            root.withdraw()
            root3 = Toplevel()
            root3.geometry("900x600")
            root3.configure(bg="pink")
            root3.resizable(False,False)
            root3.title("CALL OF DUTY: MOBILE")

            def codm_to_main():
                root3.destroy()
                root.deiconify()

            back_button = customtkinter.CTkButton(master=root3, font=('Calibri', 20, "bold"),
                                                    width=220, text="Back to Main Interface",
                                                    border_color="#e09132",
                                                    border_width=0,
                                                    hover_color="#C4661F",
                                                    fg_color="#e09132",
                                                    text_color="white",
                                                    command=lambda:codm_to_main()
                                                    )
            back_button.place(x=650,y = 10)

            excel_con=Workbook()
            excel_con=load_workbook('codm_cart_items.xlsx')
            excel_activate=excel_con.active

            cod_frame = customtkinter.CTkFrame(master=root3,
                                            height=750,
                                            width=390,
                                            bg_color="pink",
                                            fg_color="#CD1076",
                                            corner_radius=10,)
            cod_frame.place(x=1,y=1)


            chartml = ImageTk.PhotoImage(Image.open("codm_chart.jpg").resize((500, 200), Image.LANCZOS))
            chartml_l = customtkinter.CTkLabel(master=root3, image=chartml,text="")
            chartml_l.place(x=400, y=50)

            us_id = customtkinter.CTkLabel(master=root3, text="User ID:",
                                        font=("Consolas BOLD",20,"normal"),bg_color = "#CD1076")
                                        
            us_id.place(x=10,y=35)

            cod_us_id_e=customtkinter.CTkEntry(master=cod_frame, width=240,corner_radius=0,
                                        font=('Century Gothic',13), 
                                        placeholder_text=' Enter your  Ingame ID.',
                                        border_color="#a9b388",border_width=0,
                                        placeholder_text_color="#ffffff",
                                        fg_color="grey",
                                        text_color="#ffffff")
            cod_us_id_e.place(x=120,y=35)

            recharge_id = customtkinter.CTkLabel(master=cod_frame, text="Select Recharge:",
                                      font=("Consolas BOLD",20,"normal"),bg_color = "#CD1076", fg_color ="#CD1076"
                                      )
            recharge_id.place(x=10,y=90)

            price_selectt =StringVar()

            o40 = Radiobutton(master=cod_frame, text=("40 Cod Points"),bg="#CD1076",value= "40 Cod Points = 20 php", variable=price_selectt)
            o40.place(x=10,y=120)

            o100 = Radiobutton(master=cod_frame, text=("100 Cod Points"),bg="#CD1076",value= "100 Cod Points = 49 php", variable=price_selectt)
            o100.place(x=10,y=150)

            o208 = Radiobutton(master=cod_frame, text=("208 Cod Points"),bg="#CD1076",value= "208 Cod Points = 98 php", variable=price_selectt)
            o208.place(x=10,y=180)

            o328 = Radiobutton(master=cod_frame, text=("328 Cod Points"),bg="#CD1076",value= "328 Cod Points = 160 php", variable=price_selectt)
            o328.place(x=10,y=210)

            o416 = Radiobutton(master=cod_frame, text=(" 416 Cod Points"),bg="#CD1076",value= "416 Cod Points = 190 php", variable=price_selectt)
            o416.place(x=10,y=240)

            o648 = Radiobutton(master=cod_frame, text=("648 Cod Points"),bg="#CD1076",value= "648 Cod Points = 285 php", variable=price_selectt)
            o648.place(x=10,y=270)

            o1080 = Radiobutton(master=cod_frame, text=("1080 Cod Points"),bg="#CD1076",value= "1080 Cod Points = 475 php", variable=price_selectt)
            o1080.place(x=10,y=300)

            o2320 = Radiobutton(master=cod_frame, text=("2320 Cod Points"),bg="#CD1076",value= "2320 Cod Points = 945 php", variable=price_selectt)
            o2320.place(x=10,y=330)

            payment_id = customtkinter.CTkLabel(master=cod_frame, text="Select Payment:",font=("Consolas BOLD",20,"normal"))
            payment_id.place(x=10, y= 330)

            codm_payment_option = StringVar()

            load = Radiobutton(master=cod_frame, text=("Load "),bg="#CD1076", value= "Load", variable=codm_payment_option)
            load.place(x=10,y=360)

            gcash = Radiobutton(master=cod_frame, text=("Gcash "),bg="#CD1076", value= "Gcash", variable=codm_payment_option)
            gcash.place(x=10,y=390)

            paymaya = Radiobutton(master=cod_frame, text=("Pay Maya "),bg="#CD1076", value= "Pay Maya", variable=codm_payment_option)
            paymaya.place(x=130,y=360)

            Paypal = Radiobutton(master=cod_frame, text=("Paypal "),bg="#CD1076", value= "Paypal", variable=codm_payment_option)
            Paypal.place(x=130,y=390)

            bank_account = Radiobutton(master=cod_frame, text=("Bank Account "),bg="#CD1076", value= "Bank Account", variable=codm_payment_option)
            bank_account.place(x=240,y=360)

            payment_id = customtkinter.CTkLabel(master=cod_frame, text="Enter Your Payment Account Here:",font=("Consolas BOLD",15,"normal"))
            payment_id.place(x=10, y= 420)

            cod_input_payment=customtkinter.CTkEntry(master=cod_frame, width=240,corner_radius=0,
                                        font=('Century Gothic',13), 
                                        placeholder_text=' Enter your Payment Acc...',
                                        border_color="#a9b388",border_width=0,
                                        placeholder_text_color="#ffffff",
                                        fg_color="grey",
                                        text_color="#ffffff")
            cod_input_payment.place(x=10,y=450)

            excel_con=Workbook()
            excel_con=load_workbook('codm_cart_items.xlsx')
            excel_activate=excel_con.active
          
            save_button = customtkinter.CTkButton(master=cod_frame,font=('Calibri',20,"bold"),
                                            width=120, 
                                            text="add to Cart",
                                            border_color="#e09132",
                                            border_width=0,
                                            hover_color="#C4661F",
                                            fg_color="#e09132",
                                            text_color="white",
                                            command=lambda:save_function()
                                            )
            save_button.place(x=10,y=500)

            delete_button = customtkinter.CTkButton(master=cod_frame,font=('Calibri',20,"bold"),
                                            width=120, 
                                            text="Cancel Order",
                                            border_color="#e09132",
                                            border_width=0,
                                            hover_color="#C4661F",
                                            fg_color="#e09132",
                                            text_color="white",
                                            command=lambda:delete_function()
                                            )
            delete_button.place(x=150,y=500)

            codm_complete_button = customtkinter.CTkButton(master=root3, font=('Calibri', 20, "bold"),
                                                width=220, text="Complete Purchase",
                                                border_color="#e09132",
                                                border_width=0,
                                                hover_color="#C4661F",
                                                fg_color="#e09132",
                                                text_color="white",
                                                command=lambda:complete_purchase()
                                                )
            codm_complete_button.place(x=650,y=550)
        


            
            def save_function():
                cod_user = cod_us_id_e.get()
                price  = price_selectt.get()
                payment = codm_payment_option.get()
                cod_in_payment = cod_input_payment.get()

                for e_row in range(2, excel_activate.max_row + 1):
                        if cod_user == excel_activate['A' + str(e_row)].value:
                            break
                addrow = str(excel_activate.max_row +1)
                excel_activate['A' + addrow] = cod_user
                excel_activate['B' + addrow] = price
                excel_activate['C' + addrow] = payment
                excel_activate['D' + addrow] = cod_in_payment
                messagebox.showinfo("NOTIFICATION", "YOUR ORDER HAS BEEN ADDED TO CART")
                        
                refresh_data(tv1)
                excel_con.save("codm_cart_items.xlsx")

            def delete_function():
                excel_activate.delete_rows(2, excel_activate.max_row)
                excel_con.save("codm_cart_items.xlsx")
                refresh_data(tv1)
                messagebox.showinfo("NOTIFICATION", "All Order has been deleted.")

            def t_view():
                global tv1

                someframe = Frame(master=root3,width=500, height=200, bg="gray")
                someframe.place(x= 400, y= 280)

                tv1 = ttk.Treeview(someframe, height=10,) 
                yscroll = Scrollbar(someframe, orient='vertical', command=tv1.yview)
                xscroll = Scrollbar(someframe, orient='horizontal', command=tv1.xview)
                tv1.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set,)
                xscroll.pack(side ="bottom",fill ="x")
                yscroll.pack(side ="right",fill ="y")
                tv1['columns'] = ('User ID', 'Order','Payment Method','Payment acc')
                tv1.column('#0', width=0)
                tv1.column('User ID', width=120)
                tv1.column('Order', width=120)
                tv1.column('Payment Method', width=120)
                tv1.column('Payment acc', width=120)

                tv1.heading('#0', text='', anchor=W)
                tv1.heading('User ID', text='User ID', anchor=W)
                tv1.heading('Order', text='Order', anchor=CENTER)
                tv1.heading('Payment Method', text='Payment Method', anchor=W)
                tv1.heading('Payment acc', text='Payment acc', anchor=W)

                
                for e_cell in range(2,(excel_activate.max_row)+1):
                    tv1.insert(parent ='',index="end",text=str(e_cell),values=(excel_activate['A'+ str(e_cell)].value,
                                                                            excel_activate['B'+ str(e_cell)].value,
                                                                            ))
        
                ur_cart = customtkinter.CTkLabel(master=root3, text="YOUR CART:",
                                        font=("Consolas BOLD",20,"normal"),fg_color ="pink"
                                        )
                ur_cart.place(x=500,y=250)
                tv1.pack()
            t_view()

            def refresh_data(tree): 
                tree.delete(*tree.get_children())

                data = update()
                for item in data:
                    tree.insert('', 'end', values=item)
            
            def update():
                update_value = list()
                for each_cell in range(2,(excel_activate.max_row)+1):
                    update_value.append([excel_activate['A'+str(each_cell)].value,
                                        excel_activate['B'+str(each_cell)].value, 
                                        excel_activate['C'+str(each_cell)].value, 
                                        excel_activate['D'+str(each_cell)].value, ])
                return update_value
            
            def complete_purchase():
                selected_user = cod_us_id_e.get()

                if not selected_user:
                    messagebox.showerror("Error", "Please enter User ID.")
                    return
                confirmation = messagebox.askyesno("Confirmation", "Are you sure you want to complete this purchase?")

                if confirmation:
                    # Remove the user's items from the cart_items
                    for e_row in range(2, excel_activate.max_row + 1):
                        if selected_user == excel_activate['A' + str(e_row)].value:
                            excel_activate.delete_rows(e_row, e_row)
                            break

                    excel_con.save("codm_cart_items.xlsx")
                    messagebox
                    refresh_data(tv1)
                
                success_window = Toplevel()
                success_window.title("Success")
                success_window.geometry("400x400")
                
                greet_image = Image.open("greet.jpg").resize((400, 400), Image.LANCZOS)
                greet_image = ImageTk.PhotoImage(greet_image)
                greet_label = tk.Label(success_window, image=greet_image)
                greet_label.image = greet_image
                greet_label.place(x=1,y=1)
            
                ok_button = tk.Button(success_window, text="OK",font=("arial",15),bg="#e09123",fg="white", command=success_window.destroy)
                ok_button.place(x=150,y=320)

            root3.mainloop()

    fl1 = ImageTk.PhotoImage(Image.open("farlight.png").resize((150, 150), Image.LANCZOS))
    fl_l = customtkinter.CTkButton(root, image=fl1,text="",
                                     border_color="black",
                                     bg_color = "black",
                                     fg_color ="black",
                                     command=lambda:fl_interface() 
                                     
                                     )
    fl_l.place(x=650, y=230)

    def fl_interface():
            root.withdraw()
            root4 = Toplevel()
            root4.geometry("900x600")
            root4.configure(bg="pink")
            root4.resizable(False,False)
            root4.title("FARLIGHT 84")

            def fl_to_main():
                root4.destroy()
                root.deiconify()

            back_button = customtkinter.CTkButton(master=root4, font=('Calibri', 20, "bold"),
                                                    width=220, text="Back to Main Interface",
                                                    border_color="#e09132",
                                                    border_width=0,
                                                    hover_color="#C4661F",
                                                    fg_color="#e09132",
                                                    text_color="white",
                                                    command=lambda:fl_to_main()
                                                    )
            back_button.place(x=650,y = 10)


            excel_con=Workbook()
            excel_con=load_workbook('fl_cart_items.xlsx')
            excel_activate=excel_con.active

            fl_frame = customtkinter.CTkFrame(master=root4,
                                            height=750,
                                            width=390,
                                            bg_color="pink",
                                            fg_color="#CD1076",
                                            corner_radius=10,)
            fl_frame.place(x=1,y=1)


            chartml = ImageTk.PhotoImage(Image.open("farlight_chart.jpg").resize((500, 200), Image.LANCZOS))
            chartml_l = customtkinter.CTkLabel(master=root4, image=chartml,text="")
            chartml_l.place(x=400, y=50)

            us_id = customtkinter.CTkLabel(master=fl_frame, text="User ID:",
                                        font=("Consolas BOLD",20,"normal"),bg_color = "#CD1076")
                                        
            us_id.place(x=10,y=35)

            fl_us_id_e=customtkinter.CTkEntry(master=fl_frame, width=240,corner_radius=0,
                                        font=('Century Gothic',13), 
                                        placeholder_text=' Enter your  Ingame ID.',
                                        border_color="#a9b388",border_width=0,
                                        placeholder_text_color="#ffffff",
                                        fg_color="grey",
                                        text_color="#ffffff")
            fl_us_id_e.place(x=120,y=35)

            recharge_id = customtkinter.CTkLabel(master=fl_frame, text="Select Recharge:",
                                      font=("Consolas BOLD",20,"normal"),bg_color = "#CD1076", fg_color ="#CD1076"
                                      )
            recharge_id.place(x=10,y=90)

            price_selecttt =StringVar()

            fl60 = Radiobutton(master=fl_frame, text=("60 Diamonds"),bg="#CD1076",value= "60 Diamonds = 28 php", variable=price_selecttt)
            fl60.place(x=10,y=120)
            
            fl100 = Radiobutton(master=fl_frame, text=("100 Diamonds"),bg="#CD1076", value= "100 Diamonds = 45 php", variable=price_selecttt)
            fl100.place(x=10,y=150)
            
            fl165 = Radiobutton(master=fl_frame, text=("165 Diamonds"),bg="#CD1076", value= "165 Diamonds = 68 php", variable=price_selecttt)
            fl165.place(x=10,y=180)
            
            fl220 = Radiobutton(master=fl_frame, text=("220 Diamonds"),bg="#CD1076", value= "220 Diamonds = 90 php", variable=price_selecttt)
            fl220.place(x=10,y=210)
            
            fl330 = Radiobutton(master=fl_frame, text=("330 Diamonds"),bg="#CD1076", value= "330 Diamonds = 131 php", variable=price_selecttt)
            fl330.place(x=10,y=240)

            fl440 = Radiobutton(master=fl_frame, text=("440 Diamonds"),bg="#CD1076",value= "440 Diamonds = 180 php", variable=price_selecttt)
            fl440.place(x=150,y=120)
            
            fl500 = Radiobutton(master=fl_frame, text=("500 Diamonds"),bg="#CD1076", value= "500 Diamonds = 207 php", variable=price_selecttt)
            fl500.place(x=150,y=150)
            
            fl660 = Radiobutton(master=fl_frame, text=("660 Diamonds"),bg="#CD1076", value= "660 Diamonds = 262 php", variable=price_selecttt)
            fl660.place(x=150,y=180)
            
            fl880 = Radiobutton(master=fl_frame, text=("880 Diamonds"),bg="#CD1076", value= "880 Diamonds = 346 php", variable=price_selecttt)
            fl880.place(x=150,y=210)
            
            fl2240 = Radiobutton(master=fl_frame, text=("2240 Diamonds"),bg="#CD1076", value= "2240 Diamonds = 900 php", variable=price_selecttt)
            fl2240.place(x=150,y=240)

            fl4700 = Radiobutton(master=fl_frame, text=("4700 Diamonds"),bg="#CD1076", value= "4700 Diamonds = 1820 php", variable=price_selecttt)
            fl4700.place(x=80,y=270)

            payment_id = customtkinter.CTkLabel(master=fl_frame, text="Select Payment:",font=("Consolas BOLD",20,"normal"))
            payment_id.place(x=10, y= 330)
            
            fl_payment_option = StringVar()

            load = Radiobutton(master=fl_frame, text=("Load "),bg="#CD1076", value= "Load", variable=fl_payment_option)
            load.place(x=10,y=360)

            gcash = Radiobutton(master=fl_frame, text=("Gcash "),bg="#CD1076", value= "Gcash", variable=fl_payment_option)
            gcash.place(x=10,y=390)

            paymaya = Radiobutton(master=fl_frame, text=("Pay Maya "),bg="#CD1076", value= "Pay Maya", variable=fl_payment_option)
            paymaya.place(x=130,y=360)

            Paypal = Radiobutton(master=fl_frame, text=("Paypal "),bg="#CD1076", value= "Paypal", variable=fl_payment_option)
            Paypal.place(x=130,y=390)

            bank_account = Radiobutton(master=fl_frame, text=("Bank Account "),bg="#CD1076", value= "Bank Account", variable=fl_payment_option)
            bank_account.place(x=240,y=360)

            payment_id = customtkinter.CTkLabel(master=fl_frame, text="Enter Your Payment Account Here:",font=("Consolas BOLD",15,"normal"))
            payment_id.place(x=10, y= 420)

            fl_input_payment=customtkinter.CTkEntry(master=fl_frame, width=240,corner_radius=0,
                                        font=('Century Gothic',13), 
                                        placeholder_text=' Enter your Payment Acc...',
                                        border_color="#a9b388",border_width=0,
                                        placeholder_text_color="#ffffff",
                                        fg_color="grey",
                                        text_color="#ffffff")
            fl_input_payment.place(x=10,y=450)

            excel_con=Workbook()
            excel_con=load_workbook('fl_cart_items.xlsx')
            excel_activate=excel_con.active


        
            save_button = customtkinter.CTkButton(master=fl_frame,font=('Calibri',20,"bold"),
                                            width=120, 
                                            text="add to Cart",
                                            border_color="#e09132",
                                            border_width=0,
                                            hover_color="#C4661F",
                                            fg_color="#e09132",
                                            text_color="white",
                                            command=lambda:save_function()
                                            )
            save_button.place(x=10,y=500)

            delete_button = customtkinter.CTkButton(master=fl_frame,font=('Calibri',20,"bold"),
                                            width=120, 
                                            text="Cancel Order",
                                            border_color="#e09132",
                                            border_width=0,
                                            hover_color="#C4661F",
                                            fg_color="#e09132",
                                            text_color="white",
                                            command=lambda:delete_function()
                                            )
            delete_button.place(x=150,y=500)

            codm_complete_button = customtkinter.CTkButton(master=root4, font=('Calibri', 20, "bold"),
                                                width=220, text="Complete Purchase",
                                                border_color="#e09132",
                                                border_width=0,
                                                hover_color="#C4661F",
                                                fg_color="#e09132",
                                                text_color="white",
                                                command=lambda:complete_purchase()
                                                )
            codm_complete_button.place(x=650,y=550)
        


            def save_function():
                cod_user = fl_us_id_e.get()
                price  = price_selecttt.get()
                payment = fl_payment_option.get()
                cod_in_payment = fl_input_payment.get()

                for e_row in range(2, excel_activate.max_row + 1):
                        if cod_user == excel_activate['A' + str(e_row)].value:
                            break
                addrow = str(excel_activate.max_row +1)
                excel_activate['A' + addrow] = cod_user
                excel_activate['B' + addrow] = price
                excel_activate['C' + addrow] = payment
                excel_activate['D' + addrow] = cod_in_payment
                messagebox.showinfo("NOTIFICATION", "YOUR ORDER HAS BEEN ADDED TO CART")
                        
                refresh_data(tv1)
                excel_con.save("fl_cart_items.xlsx")

            def delete_function():
                excel_activate.delete_rows(2, excel_activate.max_row)
                excel_con.save("fl_cart_items.xlsx")
                refresh_data(tv1)
                messagebox.showinfo("NOTIFICATION", "All Order has been deleted.")

            
            def t_view():
                global tv1

                someframe = Frame(master=root4,width=500, height=200, bg="gray")
                someframe.place(x= 400, y= 280)

                tv1 = ttk.Treeview(someframe, height=10,) 
                yscroll = Scrollbar(someframe, orient='vertical', command=tv1.yview)
                xscroll = Scrollbar(someframe, orient='horizontal', command=tv1.xview)
                tv1.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set,)
                xscroll.pack(side ="bottom",fill ="x")
                yscroll.pack(side ="right",fill ="y")
                tv1['columns'] = ('User ID', 'Order','Payment Method','Payment acc')
                tv1.column('#0', width=0)
                tv1.column('User ID', width=120)
                tv1.column('Order', width=120)
                tv1.column('Payment Method', width=120)
                tv1.column('Payment acc', width=120)

                tv1.heading('#0', text='', anchor=W)
                tv1.heading('User ID', text='User ID', anchor=W)
                tv1.heading('Order', text='Order', anchor=CENTER)
                tv1.heading('Payment Method', text='Payment Method', anchor=W)
                tv1.heading('Payment acc', text='Payment acc', anchor=W)

                
                for e_cell in range(2,(excel_activate.max_row)+1):
                    tv1.insert(parent ='',index="end",text=str(e_cell),values=(excel_activate['A'+ str(e_cell)].value,
                                                                            excel_activate['B'+ str(e_cell)].value,
                                                                            ))
        
                ur_cart = customtkinter.CTkLabel(master=root4, text="YOUR CART:",
                                        font=("Consolas BOLD",20,"normal"),fg_color ="pink"
                                        )
                ur_cart.place(x=500,y=250)

                tv1.pack()
            t_view()

            def refresh_data(tree): 
                tree.delete(*tree.get_children())

                data = update()
                for item in data:
                    tree.insert('', 'end', values=item)
            
            def update():
                update_value = list()
                for each_cell in range(2,(excel_activate.max_row)+1):
                    update_value.append([excel_activate['A'+str(each_cell)].value,
                                        excel_activate['B'+str(each_cell)].value, 
                                        excel_activate['C'+str(each_cell)].value, 
                                        excel_activate['D'+str(each_cell)].value, ])
                return update_value

            def complete_purchase():
                selected_user = fl_us_id_e.get()

                if not selected_user:
                    messagebox.showerror("Error", "Please enter User ID.")
                    return
                confirmation = messagebox.askyesno("Confirmation", "Are you sure you want to complete this purchase?")

                if confirmation:
                    for e_row in range(2, excel_activate.max_row + 1):
                        if selected_user == excel_activate['A' + str(e_row)].value:
                            excel_activate.delete_rows(e_row, e_row)
                            break

                    excel_con.save("fl_cart_items.xlsx")
                    messagebox
                    refresh_data(tv1)
                
                success_window = Toplevel()
                success_window.title("Success")
                success_window.geometry("400x400")
                
                greet_image = Image.open("greet.jpg").resize((400, 400), Image.LANCZOS)
                greet_image = ImageTk.PhotoImage(greet_image)
                greet_label = tk.Label(success_window, image=greet_image)
                greet_label.image = greet_image
                greet_label.place(x=1,y=1)
            
                ok_button = tk.Button(success_window, text="OK",font=("arial",15),bg="#e09123",fg="white", command=success_window.destroy)
                ok_button.place(x=150,y=320)


            root4.mainloop()

Mainpage()

if __name__ == "__main__":
    Mainpage()

    
    
tk.mainloop()